import openpyxl,datetime
wb = openpyxl.Workbook()
# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = 42
# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()
# Rows can also be appended
ws.append([1, 2, 3])
# Save the file
wb.save("test2.xlsx")
